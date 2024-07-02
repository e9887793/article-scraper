from bs4 import BeautifulSoup
import seleniumwire.undetected_chromedriver as uc
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.common.action_chains import ActionChains

# Script variables
team = "pediatrics"
fileName = "_data.xlsx"
fullFile = team + fileName

def read_data(dr, url):
    dr.get(url)
    cookies = dr.get_cookies()
    # Use cookies in subsequent requests
    for cookie in cookies:
        dr.add_cookie(cookie)
    time.sleep(2)
    return dr.find_element(By.TAG_NAME, "body")
    # print(button.parent)
    # dr.implicitly_wait(2)
    # ActionChains(dr).move_to_element(button).click(button).perform()
    # dr.find_element(By.TAG_NAME, "input").click();
    # WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.CSS_SELECTOR,"iframe[title='Widget containing a Cloudflare security challenge']")))
    # WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "label.ctp-checkbox-label"))).click() 
    # wait for CloudFlare
    # wait = WebDriverWait(dr, 10)

    # wait.until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='ray-id']")))
    # now get page content
    # bs = BeautifulSoup(body.text,"html.parser")
    # text = bs.findAll(string=True)
    # return text;


def setup_driver():
    ## Set chrome Options
    options = uc.ChromeOptions()
    options.add_argument('--ignore-certificate-errors')
    # options.add_argument('--headless')
    options.add_argument("--enable-javascript")
    options.add_argument("--disable-blink-features")
    options.add_argument('--disable-blink-features=AutomationControlled')
    # options.add_argument("--start-maximized")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-setuid-sandbox")
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument("--disable-extensions")
    # options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/99.0.1234.567 Safari/537.36")

    ## Disable loading images for faster crawling
    # options.add_experimental_option('useAutomationExtension', False)
    # options.add_experimental_option("excludeSwitches", ["enable-automation"])
    

    # user_agents = [
    #     # Add your list of user agents here
    #     'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    #     'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    #     'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36',
    #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
    #     'Mozilla/5.0 (Macintosh; Intel Mac OS X 13_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/16.1 Safari/605.1.15',
    # ]

    # select random user agent
    # user_agent = random.choice(user_agents)

    # pass in selected user agent as an argument
    # options.add_argument(f'user-agent={user_agent}')

    # Initialize the WebDriver
    dr = uc.Chrome(options=options, use_subprocess=True)
    dr.maximize_window()
    return dr

def get_dss(body):
    # texts = ['Data sharing', 'Data avail']
    blocks = body.find_elements(By.TAG_NAME, "p")
    dss_list = []
    for block in blocks:
        if 'Data sharing' in block.text or 'Data avail' in block.text:
            dss_list.append(block.text)
    return dss_list

def save(text, ws, i, j):
    if len(text) != 0:
        text = "|".join(text)
        ws.cell(row = i, column = j, value=text)

def get_funding(body):
    blocks = body.find_elements(By.TAG_NAME, "p")
    fund_list = []
    for block in blocks:
        if 'funding' in block.text or 'financ' in block.text:
            fund_list.append(block.text)
    return fund_list


# def make_data_file():   
#   #don't overwrite existing file
#   df = Path(fullFile)
#   if df.is_file():
#     return
#   # doesn't exist, create it
#   headers = ['PMID', 'DOI', ]
#   wb = Workbook()
#   wb.create_sheet("data")
#   wb.save(fullFile)
#   wb.path = fullFile

# setup reader
dr = None
try:
    dr = setup_driver()
except Exception as error:
    "Error during driver setup: {0}".format(error)

wb = load_workbook(filename = fullFile)
ws = wb["data"]

# main program loop
for i in range(2, 3):
    doi = ws.cell(row = i, column = 1).value
    url = "https://doi.org/" + doi
    # url = "https://www.sciencedirect.com/science/article/pii/S0022347623004390?via%253Dihub"
    # print(url)
    body = read_data(dr, url)
    funding = get_funding(body)
    save(funding, ws, i, 2)
    dss = get_dss(body)
    save(dss, ws, i, 3)
    # save every 10 records or so
    if i % 10 == 0:
        wb.save(fullFile)
# finally save
wb.save(fullFile)